import io
import json
import csv
import textwrap
import os
from io import StringIO
from fpdf import FPDF
from config import FONT_PATH
from datetime import datetime
from openpyxl import Workbook

HISTORY_PATH = os.path.join(os.path.dirname(__file__), 'history.json')

def export_to_csv(data, description):
    """Экспорт данных в CSV формат"""
    buf = StringIO()
    if data:
        writer = csv.DictWriter(buf, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    return io.BytesIO(buf.getvalue().encode('utf-8'))


def export_to_json(data, description):
    """Экспорт данных в JSON формат"""
    output = {
        "description": description,
        "export_date": datetime.now().isoformat(),
        "data": data
    }
    return io.BytesIO(json.dumps(output, ensure_ascii=False, indent=2).encode('utf-8'))


def export_to_xlsx(data, description):
    """Экспорт данных в Excel формат"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Stats"
    
    if data:
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data[header])
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_to_pdf(data, description, filename):
    """Экспорт данных в PDF формат"""
    pdf = FPDF()
    pdf.add_page()
    
    # Добавляем шрифт для кириллицы
    if os.path.exists(FONT_PATH):
        pdf.add_font('Arial', '', FONT_PATH, uni=True)
        pdf.set_font('Arial', '', 12)
    else:
        pdf.set_font('Arial', '', 12)
    
    # Заголовок
    pdf.cell(0, 10, description, ln=True, align='C')
    pdf.cell(0, 10, f'Экспорт от {datetime.now().strftime("%d.%m.%Y %H:%M")}', ln=True, align='C')
    pdf.ln(10)
    
    if data:
        # Определяем ширину колонок
        headers = list(data[0].keys())
        col_width = 190 / len(headers)
        
        # Заголовки таблицы
        pdf.set_font('Arial', 'B', 10)
        for header in headers:
            pdf.cell(col_width, 10, str(header), border=1, align='C')
        pdf.ln()
        
        # Данные
        pdf.set_font('Arial', '', 9)
        for row in data:
            for header in headers:
                value = str(row.get(header, ''))
                # Обрезаем длинный текст
                if len(value) > 15:
                    value = value[:12] + '...'
                pdf.cell(col_width, 8, value, border=1, align='C')
            pdf.ln()
    
    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return buf


def export_to_xlsx(data, description):
    """Экспорт данных в Excel формат"""
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Stats"
        
        if data:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data[header])
        
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
    except ImportError:
        raise Exception('Ошибка экспорта в Excel. Установите: pip install openpyxl')


def export_to_pdf(data, description, filename):
    """Экспорт данных в PDF с поддержкой кириллицы через reportlab. Требует arialmt.ttf в папке с ботом."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import os

        font_name = "arialmt"
        font_path = "arialmt.ttf"
        if not os.path.exists(font_path):
            raise Exception("Для экспорта PDF с кириллицей положите arialmt.ttf в папку с ботом!")
        pdfmetrics.registerFont(TTFont(font_name, font_path))

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        styleN = styles["Normal"]
        styleH = styles["Heading1"]
        styleN.fontName = font_name
        styleH.fontName = font_name

        # Заголовок
        elements.append(Paragraph("BakS eSports - Export Data", styleH))
        elements.append(Spacer(1, 12))
        # Описание
        elements.append(Paragraph(description, styleN))
        elements.append(Spacer(1, 12))
        # Таблица
        if data:
            headers = list(data[0].keys())
            table_data = [headers] + [[str(row[h]) for h in headers] for row in data]
            t = Table(table_data, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                ('TEXTCOLOR', (0,0), (-1,0), colors.black),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTNAME', (0,0), (-1,-1), font_name),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('BOTTOMPADDING', (0,0), (-1,0), 8),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ]))
            elements.append(t)
        doc.build(elements)
        buf.seek(0)
        return buf
    except Exception as e:
        raise Exception(f'Ошибка экспорта в PDF: {str(e)}')


def export_data(data, description, filename, format_type):
    """Универсальная функция экспорта данных"""
    if format_type == 'csv':
        return export_to_csv(data, description)
    elif format_type == 'json':
        return export_to_json(data, description)
    elif format_type == 'xlsx':
        return export_to_xlsx(data, description)
    elif format_type == 'pdf':
        return export_to_pdf(data, description, filename)
    else:
        raise ValueError('Неизвестный формат экспорта')


def log_history(user_id, username, action, params=None):
    entry = {
        'user_id': user_id,
        'username': username,
        'action': action,
        'params': params or {},
        'timestamp': datetime.now().isoformat(sep=' ', timespec='seconds')
    }
    try:
        if os.path.exists(HISTORY_PATH):
            with open(HISTORY_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
        else:
            data = []
        data.append(entry)
        with open(HISTORY_PATH, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        pass